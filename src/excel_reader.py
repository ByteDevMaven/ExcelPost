import openpyxl

class ExcelReader():
    def __init__(self, path):
        wb_obj = openpyxl.load_workbook(path)
        self.sheet_obj = wb_obj.active

    def getMaxRow(self) -> int:
        return self.sheet_obj.max_row
    
    def getMaxCol(self) -> int:
        return self.sheet_obj.max_column
    
    def getRow(self, numrow: int) -> list:
        row = []

        if (numrow > self.getMaxRow()):
            numrow = self.getMaxRow()

        if (numrow < 1):
            numrow = 1

        for i in range(1, self.getMaxCol() + 1):
            cell_obj = self.sheet_obj.cell(row=numrow, column=i)
            row.append(cell_obj.value)

        return row